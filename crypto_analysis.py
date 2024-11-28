import requests
import pandas as pd
import time
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import schedule

# Fetch the top 50 cryptocurrencies by market capitalization using CoinGecko API
def fetch_crypto_data():
    url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {
        "vs_currency": "usd",
        "order": "market_cap_desc",
        "per_page": 50,
        "page": 1,
        "sparkline": False
    }
    response = requests.get(url, params=params)
    if response.status_code == 200:
        data = response.json()
        return pd.DataFrame([{
            "Cryptocurrency Name": coin["name"],
            "Symbol": coin["symbol"].upper(),
            "Current Price (USD)": coin["current_price"],
            "Market Capitalization (USD)": coin["market_cap"],
            "24-hour Trading Volume (USD)": coin["total_volume"],
            "Price Change (24h, %)": coin["price_change_percentage_24h"]
        } for coin in data])
    else:
        print("Error fetching data:", response.status_code)
        return pd.DataFrame()

# Perform basic analysis on the fetched cryptocurrency data
def analyze_data(df):
    if df.empty:
        print("No data available for analysis.")
        return None
    
    # Get the top 5 cryptocurrencies by market cap
    top_5 = df.nlargest(5, "Market Capitalization (USD)")[["Cryptocurrency Name", "Market Capitalization (USD)"]]

    # Calculate the average price of the top 50 cryptocurrencies
    avg_price = df["Current Price (USD)"].mean()

    # Find the highest and lowest 24-hour percentage price changes
    highest_change = df.nlargest(1, "Price Change (24h, %)")[["Cryptocurrency Name", "Price Change (24h, %)"]]
    lowest_change = df.nsmallest(1, "Price Change (24h, %)")[["Cryptocurrency Name", "Price Change (24h, %)"]]

    # Print the insights
    print("\n--- Analysis Report ---")
    print("Top 5 Cryptocurrencies by Market Cap:")
    print(top_5.to_string(index=False))
    print(f"\nAverage Price of Top 50 Cryptocurrencies: ${avg_price:.2f}")
    print("\nHighest 24-hour Percentage Price Change:")
    print(highest_change.to_string(index=False))
    print("\nLowest 24-hour Percentage Price Change:")
    print(lowest_change.to_string(index=False))

# Save the fetched data to an Excel sheet, updating it if the file already exists
def update_excel(df):
    file_name = "live_crypto_data.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        sheet.delete_rows(2, sheet.max_row)  # Remove old data
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(list(df.columns))  # Add headers if file doesn't exist

    # Add the new data to the sheet
    for row in dataframe_to_rows(df, index=False, header=False):
        sheet.append(row)
    
    workbook.save(file_name)
    print("Excel file updated:", file_name)

# Fetch the data, analyze it, and update the Excel sheet
def fetch_analyze_update():
    print("\nFetching live cryptocurrency data...")
    crypto_data = fetch_crypto_data()
    if not crypto_data.empty:
        analyze_data(crypto_data)
        update_excel(crypto_data)
    else:
        print("No data fetched. Skipping analysis and Excel update.")

# Schedule the task to run every 2 minutes
schedule.every(2).minutes.do(fetch_analyze_update)

# Run the initial fetch, analyze, and update
fetch_analyze_update()

# Keep the script running for continuous updates
print("\nScript is running. Press Ctrl+C to stop.")
while True:
    schedule.run_pending()
    time.sleep(1)
